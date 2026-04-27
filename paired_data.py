import pandas as pd
from scipy.stats import wilcoxon
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- 1. LOAD DATA ---
try:
    df = pd.read_csv("individual_ai_results.csv")
except FileNotFoundError:
    print("❌ Error: 'individual_ai_results.csv' not found.")
    exit()

# --- 2. CALCULATE NORMALIZED METRICS ---
df['GT_EDV'] = df['EDV'] - df['EDV_Bias']
df['GT_ESV'] = df['ESV'] - df['ESV_Bias']
df['EDV_ml'] = df['EDV_Bias'].abs()
df['ESV_ml'] = df['ESV_Bias'].abs()
df['EF_Abs'] = df['EF_Bias'].abs()
df['EDV_%'] = (df['EDV_ml'] / df['GT_EDV'].replace(0, np.nan)) * 100
df['ESV_%'] = (df['ESV_ml'] / df['GT_ESV'].replace(0, np.nan)) * 100

metrics = ['Dice', 'EF_Abs', 'EDV_ml', 'ESV_ml', 'EDV_%', 'ESV_%']

# --- 3. PAIRED COMPARISONS ---
paired_comparisons = [
    ("Age Paired", "Young Only",  "Age Balanced", {"AgeGroup": "Old"}),
    ("Age Paired", "Young Only",  "Age Balanced", {"AgeGroup": "Young"}),
    ("Age Paired", "Old Only",    "Age Balanced", {"AgeGroup": "Old"}),
    ("Age Paired", "Old Only",    "Age Balanced", {"AgeGroup": "Young"}),
    ("Sex Paired", "Female Only", "Sex Balanced", {"Sex": "M"}),
    ("Sex Paired", "Female Only", "Sex Balanced", {"Sex": "F"}),
    ("Sex Paired", "Male Only",   "Sex Balanced", {"Sex": "M"}),
    ("Sex Paired", "Male Only",   "Sex Balanced", {"Sex": "F"}),
]

def get_group_label(subgroup_filter):
    parts = []
    age = subgroup_filter.get("AgeGroup", "")
    sex = subgroup_filter.get("Sex", "")
    if age: parts.append(age)
    if sex: parts.append(sex)
    return " ".join(parts)

def filter_df(data, subgroup_filter):
    mask = pd.Series([True] * len(data), index=data.index)
    for col, val in subgroup_filter.items():
        mask &= (data[col] == val)
    return data[mask]

def run_paired(model_a_data, model_b_data, metric):
    merged = model_a_data[['sid', metric]].merge(
        model_b_data[['sid', metric]],
        on='sid',
        suffixes=('_a', '_b')
    ).dropna()
    if len(merged) < 10:
        return np.nan, np.nan, np.nan, 0
    diff = merged[f'{metric}_a'] - merged[f'{metric}_b']
    if diff.abs().sum() == 0:
        return merged[f'{metric}_a'].median(), merged[f'{metric}_b'].median(), 1.0, len(merged)
    stat, p = wilcoxon(merged[f'{metric}_a'], merged[f'{metric}_b'])
    return (
        round(merged[f'{metric}_a'].median(), 3),
        round(merged[f'{metric}_b'].median(), 3),
        p,
        len(merged)
    )

# --- 4. RUN ---
results = []
last_analysis, last_comparison = None, None

header = f"{'ANALYSIS':<22} | {'MODEL A':<15} | {'MODEL B':<15} | {'GROUP':<12} | {'METRIC':<14} | {'N':<5} | {'MED_A':<8} | {'MED_B':<8} | {'P-VALUE':<12} | {'SIG'}"
print("\n" + "="*130)
print(header)
print("="*130)

for analysis, model_a, model_b, subgroup_filter in paired_comparisons:
    group_label = get_group_label(subgroup_filter)
    comparison_key = f"{model_a}_vs_{model_b}_{group_label}"

    if last_analysis is not None and analysis != last_analysis:
        print("\n\n" + "#" * 130)
    elif last_comparison is not None and comparison_key != last_comparison:
        print("\n" + "-" * 130)
    last_analysis = analysis
    last_comparison = comparison_key

    df_a = filter_df(df[df['Task'] == model_a], subgroup_filter)
    df_b = filter_df(df[df['Task'] == model_b], subgroup_filter)

    for m in metrics:
        med_a, med_b, p, n = run_paired(df_a, df_b, m)
        if np.isnan(p):
            print(f"{analysis:<22} | {model_a:<15} | {model_b:<15} | {group_label:<12} | {m:<14} | {'<10':<5} | {'N/A':<8} | {'N/A':<8} | {'N/A':<12} | N/A")
            continue
        sig = "TRUE" if p < 0.05 else "FALSE"
        p_str = "<0.001" if p < 0.001 else f"{p:.3f}"
        print(f"{analysis:<22} | {model_a:<15} | {model_b:<15} | {group_label:<12} | {m:<14} | {n:<5} | {med_a:<8} | {med_b:<8} | {p_str:<12} | {sig}")
        results.append([analysis, model_a, model_b, group_label, m, n, med_a, med_b, p, sig])

# --- 5. BUILD EXCEL ---
FONT_NAME   = "Arial"
PINK_FILL   = PatternFill("solid", start_color="FFB3BA", end_color="FFB3BA")
GREY_FILL   = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")
HEADER_FILL = PatternFill("solid", start_color="BFBFBF", end_color="BFBFBF")
WHITE_FILL  = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
GREY_SIDE   = Side(style='thin', color='BFBFBF')
THICK_SIDE  = Side(style='medium', color='000000')
GREY_BORDER = Border(left=GREY_SIDE, right=GREY_SIDE, top=GREY_SIDE, bottom=GREY_SIDE)

def make_font(bold=False):
    return Font(name=FONT_NAME, bold=bold, size=10)

wb = Workbook()
ws = wb.active
ws.title = "Paired Results"

headers = ["Analysis", "Model_A", "Model_B", "Group", "Metric", "N", "Med_A", "Med_B", "P-Value", "Significant"]
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = make_font(bold=True)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = GREY_BORDER

for i in range(1, 11):
    ws.column_dimensions[get_column_letter(i)].width = 14

current_row = 2
prev_analysis = None
i = 0

while i < len(results):
    key = (results[i][0], results[i][1], results[i][2], results[i][3])
    group_rows = []
    j = i
    while j < len(results) and (results[j][0], results[j][1], results[j][2], results[j][3]) == key:
        group_rows.append(results[j])
        j += 1

    analysis, model_a, model_b, group = key
    start_row = current_row
    end_row = current_row + len(group_rows) - 1

    if prev_analysis is not None and analysis != prev_analysis:
        for col in range(1, 11):
            c = ws.cell(row=current_row, column=col)
            c.border = Border(top=THICK_SIDE, left=GREY_SIDE, right=GREY_SIDE, bottom=GREY_SIDE)

    prev_analysis = analysis

    for idx, row_data in enumerate(group_rows):
        excel_row = current_row + idx
        is_sig = row_data[9] == "TRUE"
        fill = PINK_FILL if is_sig else WHITE_FILL

        # Col A — grey always
        a_cell = ws.cell(row=excel_row, column=1, value=analysis if idx == 0 else None)
        a_cell.fill = GREY_FILL
        a_cell.font = make_font(bold=True)
        a_cell.alignment = Alignment(horizontal='center', vertical='center')
        a_cell.border = GREY_BORDER

        # Cols B, C, D — white
        for col_idx, val in zip([2, 3, 4], [model_a, model_b, group]):
            cell = ws.cell(row=excel_row, column=col_idx, value=val if idx == 0 else None)
            cell.fill = WHITE_FILL
            cell.font = make_font(bold=True if idx == 0 else False)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = GREY_BORDER

        # Cols E-H: Metric, N, Med_A, Med_B
        for col_idx, val in enumerate(row_data[4:8], 5):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font = make_font(bold=is_sig)
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = GREY_BORDER

        # Col I — P-Value: <0.001 as string, else 3dp number
        p_val = row_data[8]
        if p_val < 0.001:
            p_cell = ws.cell(row=excel_row, column=9, value="<0.001")
        else:
            p_cell = ws.cell(row=excel_row, column=9, value=round(p_val, 3))
            p_cell.number_format = '0.000'
        p_cell.font = make_font(bold=is_sig)
        p_cell.fill = fill
        p_cell.alignment = Alignment(horizontal='center', vertical='center')
        p_cell.border = GREY_BORDER

        # Col J — Significant
        sig_cell = ws.cell(row=excel_row, column=10, value=row_data[9])
        sig_cell.font = make_font(bold=is_sig)
        sig_cell.fill = fill
        sig_cell.alignment = Alignment(horizontal='center', vertical='center')
        sig_cell.border = GREY_BORDER

        ws.row_dimensions[excel_row].height = 18

    if len(group_rows) > 1:
        for col in range(1, 5):
            ws.merge_cells(start_row=start_row, start_column=col,
                           end_row=end_row, end_column=col)
            cell = ws.cell(row=start_row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = GREY_BORDER
            cell.fill = GREY_FILL if col == 1 else WHITE_FILL
            cell.font = make_font(bold=True)

    current_row += len(group_rows)
    i = j

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 18

output_file = "paired_disparity_results.xlsx"
wb.save(output_file)
print(f"\n✅ Done! Paired Excel saved to: {output_file}")